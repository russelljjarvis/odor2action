"""
Author: [Russell Jarvis](https://github.com/russelljjarvis)

"""
import matplotlib
matplotlib.use("Agg")
from matplotlib.backends.backend_agg import RendererAgg
_lock = RendererAgg.lock

import seaborn as sns

from community import community_louvain
#import igraph as ig
import plotly.graph_objs as go
from matplotlib.patches import FancyArrowPatch, Circle
import numpy as np
import holoviews as hv
from holoviews import opts, dim
from bokeh.plotting import show, output_file
import copy

import argparse
import numpy as np
import networkx as nx
import streamlit as st

import streamlit.components.v1 as components
import networkx as nx
import matplotlib.pyplot as plt
from pyvis.network import Network
import shelve

# import streamlit as st
import os
import pandas as pd
import pickle
import streamlit as st

# from holoviews import opts, dim
from collections import Iterable
import networkx

# import holoviews as hv
# import chord2
import shelve

import plotly.graph_objects as go
import pandas as pd
import streamlit as st
import numpy as np
import pickle

import plotly.graph_objects as go
import tkinter
#import streamlit as st

#def main():


import pandas as pd
import openpyxl
from pathlib import Path
import numpy as np
import networkx as nx

import xlrd
import matplotlib.pyplot as plt
from community import community_louvain


import numpy as np
import matplotlib.pyplot as plt
import networkx as nx
from pyveplot import Hiveplot, Axis, Node
import networkx as nx
import random
import base64
import textwrap




def disable_logo(plot, element):
    plot.state.toolbar.logo = None


# hv.extension("bokeh", logo=False)
# hv.output(size=150)
# hv.plotting.bokeh.ElementPlot.finalize_hooks.append(disable_logo)

pd.set_option("display.max_rows", None)
pd.set_option("display.max_columns", None)
pd.set_option("display.width", None)
pd.set_option("display.max_colwidth", -1)
import plotly.graph_objects as go

# from auxillary_methods import plotly_sized2

from datashader.bundling import hammer_bundle

from typing import List
import pandas as pd

# import holoviews as hv


def generate_sankey_figure(
    nodes_list: List, edges_df: pd.DataFrame, title: str = "Sankey Diagram"
):

    edges_df["src"] = edges_df["src"].apply(lambda x: nodes_list.index(x))
    edges_df["tgt"] = edges_df["tgt"].apply(lambda x: nodes_list.index(x))
    # creating the sankey diagram
    data = dict(
        type="sankey",
        node=dict(
            hoverinfo="all",
            pad=15,
            thickness=20,
            line=dict(color="black", width=0.5),
            label=nodes_list,
        ),
        link=dict(
            source=edges_df["src"], target=edges_df["tgt"], value=edges_df["weight"]
        ),
    )

    layout = dict(title=title, font=dict(size=10))

    fig = go.Figure(data=[data], layout=layout)
    st.write(fig)

    # return fig


# def fix():
# 02P1 missing
# 13P1 IRG3 leader. Currently Blue should be green.
# move from IRG1 to IRG3

# @st.cache
# @st.cache(suppress_st_warning=True)
def data_shade(graph, color_code, adj_mat, color_dict, labels_=False):

    nodes = graph.nodes
    # orig_pos=nx.get_node_attributes(graph,'pos')

    nodes_ind = [i for i in range(0, len(graph.nodes()))]
    redo = {k: v for k, v in zip(graph.nodes, nodes_ind)}
    # pos = nx.spring_layout(H, k=0.05, seed=4572321, scale=1)

    pos_ = nx.spring_layout(graph, scale=2.5, k=0.00015, seed=4572321)
    # node_color = [community_index[n] for n in graph]
    H = graph.to_undirected()
    centrality = nx.betweenness_centrality(H, k=10, endpoints=True)

    node_size = [v * 25000 for v in centrality.values()]

    coords = []
    for node in graph.nodes:
        x, y = pos_[node]
        coords.append((x, y))
    nodes_py = [
        [new_name, pos[0], pos[1]]
        for name, pos, new_name in zip(nodes, coords, nodes_ind)
    ]
    ds_nodes = pd.DataFrame(nodes_py, columns=["name", "x", "y"])
    ds_edges_py = []
    for (n0, n1) in graph.edges:
        ds_edges_py.append([redo[n0], redo[n1]])
    ds_edges = pd.DataFrame(ds_edges_py, columns=["source", "target"])
    hb = hammer_bundle(ds_nodes, ds_edges)
    hbnp = hb.to_numpy()
    splits = (np.isnan(hbnp[:, 0])).nonzero()[0]
    start = 0
    segments = []
    # st.markdown(
    #    'It is also a bit like how parallel neurons are wrapped together closely by a density of myline cells, like neurons traveling through the corpus callosum'
    # )
    # st.markdown(
    #    "Think of it conceptually like Ramon Y Cajal principle of wiring cost optimization."
    # )
    # st.markdown(
    #    "Neurons processes projecting to very close places shouldnt travel down indipendant dedicated lines, there is less metabolic cost involved in \n channeling parallel fibres in the same myline sheath through a backbone like the corpus callosum."
    # )

    for stop in splits:
        seg = hbnp[start:stop, :]
        segments.append(seg)
        start = stop

    fig, ax = plt.subplots(figsize=(15, 15))
    widths = list(adj_mat["weight"].values)
    srcs = list(adj_mat["src"].values)

    for ind, seg in enumerate(segments):
        ax.plot(
            seg[:, 0],
            seg[:, 1],
            c=color_code[srcs[ind]],
            alpha=0.67,
            linewidth=widths[ind],
        )
    node_color = [color_code[n] for n in graph]

    ax3 = nx.draw_networkx_nodes(
        graph,
        pos_,
        node_color=node_color,
        node_size=node_size,
        node_shape="o",
        alpha=0.5,
        vmin=None,
        vmax=None,
        linewidths=2.0,
        label=None,
        ax=ax,
    )  # , **kwds)

    axx = plt.gca()  # to get the current axis

    axx.collections[0].set_edgecolor("#FF0000")
    labels = {}
    for node in graph.nodes():
        # set the node name as the key and the label as its value
        labels[node] = node
    if labels_:
        nx.draw_networkx_labels(graph, pos_, labels, font_size=16, font_color="r")

    # ax3.margins(0.1, 0.05)
    fig.tight_layout()
    plt.axis("off")
    for k, v in color_dict.items():
        plt.scatter([], [], c=v, label=k)
    plt.legend(frameon=False, prop={"size": 24})

    def dontdo(segments, pos_, graph):
        fig.show()
        df_geo = {}
        df_geo["text"] = list(node for node in graph.nodes)

        fig = go.Figure()
        lats = []
        lons = []
        traces = []
        other_traces = []
        # from tqdm import tqdm
        for ind, seg in enumerate(segments):
            x0, y0 = seg[1, 0], seg[1, 1]  # graph.nodes[edge[0]]['pos']
            x1, y1 = seg[-1, 0], seg[-1, 1]  # graph.nodes[edge[1]]['pos']
            xx = seg[:, 0]
            yy = seg[:, 1]
            lats.append(xx)
            lons.append(yy)
            for i, j in enumerate(xx):
                if i > 0:
                    other_traces.append(
                        go.Scatter(
                            lon=[xx[i], xx[i - 1]],
                            lat=[yy[i], yy[i - 1]],
                            mode="lines",
                            showlegend=False,
                            hoverinfo="skip",
                            line=dict(width=0.5, color="blue"),
                        )
                    )
        fig.add_traces(other_traces)
        fig.add_trace(
            go.Scatter(
                lat=pos_,
                lon=pos_,
                marker=dict(
                    size=3,  # data['Confirmed-ref'],
                    color=colors,
                    opacity=1,
                ),
                text=list(graph.nodes),
                hovertemplate="%{text} <extra></extra>",
            )
        )
        fig["layout"]["width"] = 1825
        fig["layout"]["height"] = 1825
        st.write(fig)

    return fig  # ,colors

    # return fig


def depricated():
    def plot_stuff(df2, edges_df_full, first, adj_mat_dicts):
        with shelve.open("fast_graphs_splash.p") as db:
            flag = "chord" in db
            if False:  # flag:
                graph = db["graph"]

            else:
                db.close()


# from hiveplotlib import Axis, Node, HivePlot
# from hiveplotlib.viz import axes_viz_mpl, node_viz_mpl, edge_viz_mpl
from matplotlib.lines import Line2D
import matplotlib.pyplot as plt


class renamer:
    def __init__(self):
        self.d = dict()

    def __call__(self, x):
        if x not in self.d:
            self.d[x] = 0
            return x
        else:
            self.d[x] += 1
            return "%s_%d" % (x, self.d[x])


def dontdo():
    df2.rename(columns=renamer(), inplace=True)

    df4 = pd.DataFrame()

    for col in df2.columns[0 : int(len(df2.columns) / 2)]:
        # if
        if col + str("_1") in df2.columns:
            df4[col] = df2[col] + df2[col + str("_1")]
        else:
            df4[col] = df2[col]


def dontdo():
    """
    for col in df2.columns:

                    if col.split("_")[0] in df3.columns and col.split("_")[0] in df5.columns:
                                    df4[col] = df2[col] + df3[col]
                                    #st.text('yes')

                    else:
                                    df4[col] = df2[col]
    """


import copy

# @st.cache(persist=True)
# @st.cache(allow_output_mutation=True)
def get_frame(transpose=False, threshold=6):

    hard_codes = Path("code_by_IRG.xlsx")
    hard_codes = openpyxl.load_workbook(hard_codes)

    hard_codes = hard_codes.active

    hard_codes = pd.DataFrame(hard_codes.values)

    xlsx_file0 = Path("o2anetmap2021.xlsx")
    xlsx_file1 = Path("o2anetmap.xlsx")
    wb_obj0 = openpyxl.load_workbook(xlsx_file0)
    wb_obj1 = openpyxl.load_workbook(xlsx_file1)

    # Read the active sheet:
    worksheet0 = wb_obj0.active
    worksheet1 = wb_obj1.active

    df3 = pd.DataFrame(worksheet0.values)
    df3.replace("", "Barely or never", regex=True, inplace=True)
    df2 = pd.DataFrame(worksheet1.values)
    df2.replace("", "Barely or never", regex=True, inplace=True)
    df3.drop(0, inplace=True)
    #df3.rename
    to_rename = {1:43,2:44,3:45,4:46}# v for k, v in zip(row_names, names)}
    df3.rename(index=to_rename, inplace=True)

    df2 = pd.concat([df2, df3], axis=0)  # ,inplace=True)
    # st.write(df2)

    sheet = copy.copy(df2)
    hc = {
        k: str("IRG ") + str(v) for k, v in zip(hard_codes[0][1::], hard_codes[1][1::])
    }
    hc["13P1"] = "IRG 3"
    hc1 = {k: "DCMT" for k, v in hc.items() if v == "IRG DCMT"}

    hc.update(hc1)
    hc.pop("Code", None)
    color_code_0 = {k: v for k, v in zip(df2[0], df2[1]) if k not in "Rater Code"}
    color_code_0.update(hc)

    # Ribbon color code needs to labeled as to or from.
    # source or target.

    color_dict = {
        "IRG 1": "blue",
        "IRG 2": "red",
        "IRG 3": "green",
        "DCMT": "purple",
    }
    color_code_1 = {}

    popg = nx.DiGraph()

    for k, v in color_code_0.items():

        if v not in popg.nodes:
            popg.add_node(v, name=v)
        color_code_1[k] = color_dict[v]
    col_to_rename = df2.columns
    ratercodes = df2[0][1::]
    row_names = list(df2.T[0].values)
    row_names.append(list(df2.T[0].values)[-1])
    row_names = row_names[2::]
    # st.text(row_names)
    names = [rn.split("- ") for rn in row_names]
    # st.text(names)

    names2 = []
    for i in names:
        if len(i) == 2:
            names2.append(i[1])
        else:
            names2.append(i)
    names = names2

    r_names = list(df2.index.values[:])

    to_rename_ind = {v: k for k, v in zip(df2[0], r_names)}

    row_names = list(range(0, len(df2.columns) + 1, 1))
    to_rename = {k: v for k, v in zip(row_names, names)}
    to_rename[113] = "12P2"

    del df2[0]
    del df2[1]
    df2.drop(0, inplace=True)
    df2.rename(index=to_rename_ind, inplace=True)
    df2.rename(columns=to_rename, inplace=True)

    unk = []

    for col in df2.columns:
        if col in df2.index.values[:]:
            pass
        else:
            pass

    legend = {}

    legend.update({"Never": 0.0})
    legend.update({"Barely or never": 1})
    legend.update({"Occasionally in a minor way": 2})
    legend.update({"Less than once a month": 3})
    legend.update({"More than once a month (But not weekly)": 4})
    legend.update({"Occasionally but substantively": 5})
    legend.update({"More than twice a week": 6})
    legend.update({"Often": 7})
    legend.update({"Much or all of the time": 8})
    legend.update({"1-2 times a week": 9.0})
    df2.replace({"": 0.0}, inplace=True)
    df2.replace({" ": 0.0}, inplace=True)
    df2.replace({"\t": 0.0}, inplace=True)
    df2.replace({"\n": 0.0}, inplace=True)

    df2.replace({"Never": 0.0}, inplace=True)
    df2.replace({"Barely or never": 1}, inplace=True)
    df2.replace({"Occasionally in a minor way": 2}, inplace=True)
    df2.replace({"Less than once a month": 3}, inplace=True)
    df2.replace({"More than once a month (But not weekly)": 4}, inplace=True)
    df2.replace({"Occasionally but substantively": 5}, inplace=True)
    df2.replace({"More than twice a week": 6}, inplace=True)
    df2.replace({"Often": 7}, inplace=True)
    df2.replace({"Much or all of the time": 8}, inplace=True)
    df2.replace({"1-2 times a week": 9.0}, inplace=True)

    # This sums columns under the same name
    df2 = df2.groupby(df2.columns, axis=1).sum()
    df2 = df2.groupby(level=0, axis=1).sum()
    df2 = df2.T
    df2 = df2.groupby(df2.columns, axis=1).sum()
    df2 = df2.groupby(level=0, axis=1).sum()
    df2 = df2.T

    if transpose:
        df2 = df2.T
    # st.write(df2["02P1"])
    # df2 = df4
    # store["df2"] = df2  # save it
    # st.write(df2)
    # store["names"] = names  # save it
    # store["ratercodes"] = ratercodes  # save it
    # store["legend"] = legend  # save it

    return (
        df2,
        names,
        ratercodes,
        legend,
        color_code_1,
        color_dict,
        color_code_0,
        sheet,
        popg,
        hc,
    )


sns_colorscale = [
    [0.0, "#3f7f93"],  # cmap = sns.diverging_palette(220, 10, as_cmap = True)
    [0.071, "#5890a1"],
    [0.143, "#72a1b0"],
    [0.214, "#8cb3bf"],
    [0.286, "#a7c5cf"],
    [0.357, "#c0d6dd"],
    [0.429, "#dae8ec"],
    [0.5, "#f2f2f2"],
    [0.571, "#f7d7d9"],
    [0.643, "#f2bcc0"],
    [0.714, "#eda3a9"],
    [0.786, "#e8888f"],
    [0.857, "#e36e76"],
    [0.929, "#de535e"],
    [1.0, "#d93a46"],
]


# @st.cache
def df_to_plotly(df, log=False):
    return {"z": df.values.tolist(), "x": df.columns.tolist(), "y": df.index.tolist()}


# @st.cache
def plot_df_plotly(sleep_df):
    fig = go.Figure(data=go.Heatmap(df_to_plotly(sleep_df, log=True)))
    st.write(fig)


# @st.cache
def plot_imshow_plotly(sleep_df):

    heat = go.Heatmap(df_to_plotly(sleep_df), colorscale=sns_colorscale)
    title = "Adjacency Matrix"

    layout = go.Layout(
        title_text=title,
        title_x=0.5,
        width=600,
        height=600,
        xaxis_showgrid=False,
        yaxis_showgrid=False,
        yaxis_autorange="reversed",
    )

    fig = go.Figure(data=[heat], layout=layout)

    st.write(fig)


def learn_embeddings(walks):
    """
    Learn embeddings by optimizing the Skipgram objective using SGD.
    """
    walks = [map(str, walk) for walk in walks]
    model = Word2Vec(
        walks,
        size=args.dimensions,
        window=args.window_size,
        min_count=0,
        sg=1,
        workers=args.workers,
        iter=args.iter,
    )
    model.save_word2vec_format(args.output)

    return


# @st.cache(persist=True)
def get_table_download_link_csv(df):
    import base64

    # csv = df.to_csv(index=False)
    csv = df.to_csv().encode()
    # b64 = base64.b64encode(csv.encode()).decode()
    b64 = base64.b64encode(csv).decode()
    href = f'<a href="data:file/csv;base64,{b64}" download="captura.csv" target="_blank">Download csv file</a>'
    return href


def draw_network(G, pos, ax, widths, edge_colors, sg=None):

    for n in G.nodes:
        c = Circle(pos[n], radius=0.05, alpha=0.7)
        # ax.add_patch(c)
        G.nodes[n]["patch"] = c
        x, y = pos[n]
    seen = {}
    for n, (u, v, d) in enumerate(G.edges(data=True)):
        n1 = G.nodes[u]["patch"]
        n2 = G.nodes[v]["patch"]
        rad = 0.1
        if (u, v) in seen:
            rad = seen.get((u, v))
            rad = (rad + np.sign(rad) * 0.1) * -1
        alpha = 0.5
        color = "k"

        e = FancyArrowPatch(
            n1.center,
            n2.center,
            patchA=n1,
            patchB=n2,
            arrowstyle="-|>",
            connectionstyle="arc3,rad=%s" % rad,
            mutation_scale=10.0,
            lw=widths[n],
            alpha=alpha,
            color=edge_colors[n],
        )
        seen[(u, v)] = rad
        ax.add_patch(e)
    return e


# @st.cache(allow_output_mutation=True,suppress_st_warning=True)
def population(cc, popg, color_dict):
    with _lock:

        fig, ax = plt.subplots(figsize=(20, 15))

        pos = nx.spring_layout(popg, k=15, seed=4572321, scale=1.5)
        sizes = {}
        for k, v in cc.items():
            if v not in sizes.keys():
                sizes[v] = 1
            else:
                sizes[v] += 1
        temp = list([s * 1000 for s in sizes.values()])
        node_color = [color_dict[n] for n in popg]
        nx.draw_networkx_nodes(
            popg,
            pos=pos,
            node_color=node_color,  # = [color_code[n] for n in H],
            node_size=temp,
            alpha=0.6,
            linewidths=2,
        )
        # if labelsx:
        #    labels = {}
        #    for node in temp.nodes():
        # set the node name as the key and the label as its value
        #        labels[node] = node
        #    nx.draw_networkx_labels(temp, label_pos, labels, font_size=29.5, font_color="b")

        widths = []  # [e["weight"] for e in popg.edges]
        # st.text(widths)
        edge_list = []
        edge_colors = []
        for e in popg.edges:
            edge_list.append((e[0], e[1]))
            edge_colors.append(color_dict[e[0]])

            ee = popg.get_edge_data(e[0], e[1])
            widths.append(ee["weight"] * 0.02)

        ax = plt.gca()
        draw_network(popg, pos, ax, widths, edge_colors)
        # labels = {v.name:v for v,v in popg.nodes}
        labels = {}
        for node in popg.nodes():
            # set the node name as the key and the label as its value
            labels[node] = node

        # for k, v in labels.items():
        #    plt.scatter([], [], c=color_dict[v], label=k)
        # plt.legend(frameon=False, prop={"size": 34})
        ax.margins(0.1, 0.1)

        popgc = copy.copy(popg)
        ax.autoscale()
        plt.axis("equal")
        plt.axis("off")
        plt.tight_layout()

        st.pyplot(fig, use_column_width=True)


def interactive_population(cc, popg, color_dict):

    sizes = {}
    for k, v in cc.items():
        if v not in sizes.keys():
            sizes[v] = 1
        else:
            sizes[v] += 1
    temp = list([s * 1000 for s in sizes.values()])
    node_color = [color_dict[n] for n in popg]

    widths = []  # [e["weight"] for e in popg.edges]
    # st.text(widths)
    edge_list = []
    edge_colors = []
    for e in popg.edges:
        edge_list.append((e[0], e[1]))
        edge_colors.append(color_dict[e[0]])

        ee = popg.get_edge_data(e[0], e[1])
        widths.append(ee["weight"] * 0.02)
    labels = {}
    for node in popg.nodes():
        # set the node name as the key and the label as its value
        labels[node] = node
    nt = Network("700px", "700px",directed=True)  # ,layout=physics_layouts)

    nt.barnes_hut()
    for node in popg.nodes:
        nt.add_node(node,label=node) # node id = 1 and label = Node 1
    nt.set_edge_smooth('continuous')

    for e in popg.edges:
        src = e[0]
        dst = e[1]
        src = str(src)
        dst = str(dst)

        ee = popg.get_edge_data(e[0], e[1])

        nt.add_edge(src, dst, width=0.01*ee["weight"])#, arrowStrikethrough=True)
    for i,node in enumerate(nt.nodes):
        node["size"] = sizes[node["id"]] #* 1025
        node["color"] = node_color[i]

    for node in nt.nodes:
        node["title"] = (
            "<br> {0}'s' group size is: {1}<br>".format(node["id"],sizes[node["id"]])
        )

    nt.save_graph("population.html")
    # nt.save_graph("saved_html.html")
    HtmlFile = open("population.html", "r", encoding="utf-8")
    source_code = HtmlFile.read()
    components.html(source_code, height=1200, width=1000)


def community_layout(g, partition):
    """
    Compute the layout for a modular graph.
    Arguments:
    ----------
    g -- networkx.Graph or networkx.DiGraph instance
        graph to plot
    partition -- dict mapping int node -> int community
        graph partitions
    Returns:
    --------
    pos -- dict mapping int node -> (float x, float y)
        node positions
    """
    pos_communities = _position_communities(
        g, partition, k=0.04, scale=5.0, seed=4572321
    )
    pos_nodes = _position_nodes(g, partition, k=0.04, scale=1.0, seed=4572321)
    # combine positions
    pos = dict()
    for node in g.nodes():
        pos[node] = pos_communities[node] + pos_nodes[node]

    return pos, pos_communities


def _position_communities(g, partition, **kwargs):
    # create a weighted graph, in which each node corresponds to a community,
    # and each edge weight to the number of edges between communities
    between_community_edges = _find_between_community_edges(g, partition)
    communities = set(partition.values())
    hypergraph = nx.DiGraph()
    hypergraph.add_nodes_from(communities)
    for (ci, cj), edges in between_community_edges.items():
        hypergraph.add_edge(ci, cj, weight=len(edges))

    # find layout for communities
    pos_communities = nx.spring_layout(hypergraph, **kwargs)

    # set node positions to position of community
    pos = dict()
    for node, community in partition.items():
        pos[node] = pos_communities[community]

    return pos


def _find_between_community_edges(g, partition):

    edges = dict()

    for (ni, nj) in g.edges():
        ci = partition[ni]
        cj = partition[nj]

        if ci != cj:
            try:
                edges[(ci, cj)] += [(ni, nj)]
            except KeyError:
                edges[(ci, cj)] = [(ni, nj)]

    return edges


def _position_nodes(g, partition, **kwargs):
    """
    Positions nodes within communities.
    """

    communities = dict()
    for node, community in partition.items():
        try:
            communities[community] += [node]
        except KeyError:
            communities[community] = [node]

    pos = dict()
    for ci, nodes in communities.items():
        subgraph = g.subgraph(nodes)
        pos_subgraph = nx.spring_layout(subgraph, **kwargs)
        pos.update(pos_subgraph)

    return pos


def dont():
    """
    def colored_hive_axis(first,color_code_0,reverse):
        c = ['#e41a1c', '#377eb8', '#4daf4a',
             '#984ea3', '#ff7f00', '#ffff33',
             '#a65628', '#f781bf', '#999999',]


        # create hiveplot object
        h = Hiveplot()



        fig = plt.figure()
        # create three axes, spaced at 120 degrees from each other

        h.axes = [Axis(start=20, angle=0,
                       stroke=random.choice(c), stroke_width=1.1),
                  Axis(start=20, angle=90,
                       stroke=random.choice(c), stroke_width=1.1),
                  Axis(start=20, angle=90 + 90,
                       stroke=random.choice(c), stroke_width=1.1)]


                  #Axis(start=20, angle=90 + 90 + 90,
                  #        stroke=random.choice(c), stroke_width=1.1)
                  #]

        fig = plt.figure()
        # create three axes, spaced at 120 degrees from each other
        h.axes = [Axis(start=20, angle=0,
                       stroke=random.choice(c), stroke_width=1.1),
                  Axis(start=20, angle=120,
                       stroke=random.choice(c), stroke_width=1.1),
                  Axis(start=20, angle=120 + 120,
                       stroke=random.choice(c), stroke_width=1.1)
                  ]

        #g = first

        # place these nodes into our three axes
        for axis, nodes in zip(h.axes,
                               [IRG1_indices, IRG2_indices, IRG3_indices]):
            circle_color = random.choice(c)
            for v in nodes:
                st.text(v)
                # create node object
                node = Node(radius=15,
                            label="node %s" % v)
                # add it to axis
                st.text(node)
                axis.add_node(v, node)
                # once it has x, y coordinates, add a circle
                node.add_circle(fill=circle_color, stroke=circle_color,
                                stroke_width=0.1, fill_opacity=0.7)
                if axis.angle < 180:
                    orientation = -1
                    scale = 0.6
                else:
                    orientation = 1
                    scale = 0.35
                # also add a label
                node.add_label("node %s" % (v),
                               angle=axis.angle + 90 * orientation,
                               scale=scale)

        # iterate through axes, from left to right
        for n in range(-1, len(h.axes) - 1):
            curve_color = random.choice(c)
            # draw curves between nodes connected by edges in network
            h.connect_axes(h.axes[n],
                           h.axes[n+1],
                           g.edges,
                           stroke_width=0.5,
                           stroke=curve_color)
        # save output
        h.save('col_ba_hiveplot1.svg')
        #from PIL import Image
        f = open('col_ba_hiveplot1.svg',"r")
        lines = f.readlines()
        line_string=''.join(lines)

        render_svg(line_string)


        #st.image(Image.open("col_ba_hiveplot.svg"))
    """


import matplotlib.patches as patches
from community import community_louvain

# @st.cache(allow_output_mutation=True,suppress_st_warning=True)
def community(first, color_code, color_dict):
    colors = [
        "#e41a1c",
        "#377eb8",
        "#4daf4a",
        "#984ea3",
        "#ff7f00",
        "#ffff33",
        "#a65628",
        "#f781bf",
        "#999999",
    ]

    my_expander = st.expander("Toggle node labels")
    labels_ = my_expander.radio("Would you like to label nodes?", ("No", "Yes"))
    if labels_ == "Yes":
        labelsx = True
    if labels_ == "No":
        labelsx = False

    try:
        with open("positions.p", "rb") as f:
            [pos, pos_communities, partition, temp] = pickle.load(f)  # ,)
    except:
        temp = first.to_undirected()
        partition = community_louvain.best_partition(temp, resolution=4.0)
        pos, pos_communities = community_layout(temp, partition)
        with open("positions.p", "wb") as f:
            pickle.dump([pos, pos_communities, partition, temp], f)
    diffcc = list(partition.values())
    pkeys = set(partition.values())
    partitiondf = pd.DataFrame([partition]).T
    pointss = []
    whichpkeys = []
    centrex = []
    centrey = []

    for k in pkeys:  # iterate over communities
        list_of_nodes = partitiondf[partitiondf.values == k].index.values[:]
        tocvxh = []
        for node in list_of_nodes:
            x, y = pos[node]
            tocvxh.append((x, y))

        meanx = np.mean([i[0] for i in tocvxh])
        meany = np.mean([i[1] for i in tocvxh])
        centrex.append(meanx)
        centrey.append(meany)

    srcs = []
    widths = []

    for e in temp.edges:
        src = partition[e[0]]
        srcs.append(src)
        ee = temp.get_edge_data(e[0], e[1])
        widths.append(1.85 * ee["weight"])
    with _lock:

        fig1, ax1 = plt.subplots(1, 1, figsize=(20, 20))
        recolored = [colors[i] for i in diffcc]

        nx.draw_networkx_nodes(
            temp,
            pos=pos,
            node_color=recolored,
            node_size=550,
            alpha=0.5,
            linewidths=1,
        )
        # axx = ax1.gca()  # to get the current axis
        # axx.collections[0].set_edgecolor("#FF0000")
        label_pos = copy.deepcopy(pos)
        for k, v in label_pos.items():
            label_pos[k][0] = v[0] + 0.5

        if labelsx:
            labels = {}
            for node in temp.nodes():
                # set the node name as the key and the label as its value
                labels[node] = node
            nx.draw_networkx_labels(temp, label_pos, labels, font_size=29.5, font_color="b")

        nx.draw_networkx_edges(temp, pos=pos, edge_color="grey", alpha=0.15, width=widths)
        for centre in zip(centrex, centrey, pkeys):
            r = 1.5
            c = (float(centre[0]), float(centre[1]))
            ax1.add_patch(plt.Circle(c, r, color=colors[centre[2]], alpha=0.15))

        def last_fig(
            color_code, first, temp, pos, pkeys, centrex, centrey, color_dict, labelsx=False
        ):
            fig2, ax2 = plt.subplots(1, 1, figsize=(20, 20))

            node_color = [color_code[n] for n in temp]
            srcs = []
            for e in temp.edges:
                src = color_code[e[0]]
                srcs.append(src)

            nx.draw_networkx_nodes(
                temp,
                pos=pos,
                node_color=node_color,
                node_size=550,
                alpha=0.5,
                linewidths=1,
            )
            label_pos = copy.deepcopy(pos)
            for k, v in label_pos.items():
                label_pos[k][0] = v[0] + 0.5

            if labelsx:
                labels = {}
                for node in temp.nodes():
                    # set the node name as the key and the label as its value
                    labels[node] = node
                nx.draw_networkx_labels(
                    temp, label_pos, labels, font_size=29.5, font_color="b"
                )

            # axx = ax2.gca()  # to get the current axis
            # axx.collections[0].set_edgecolor("#FF0000")
            nx.draw_networkx_edges(
                temp, pos=pos, edge_color="grey", alpha=0.15, width=widths
            )
            for centre in zip(centrex, centrey, pkeys):
                r = 1.5
                c = (float(centre[0]), float(centre[1]))
                ax2.add_patch(plt.Circle(c, r, color=colors[centre[2]], alpha=0.15))
                #    if labelsx:
                # for k, v in color_dict.items():
                #    plt.scatter([], [], c=v, label=k)
                # plt.legend(frameon=False, prop={"size": 29.5})
            return fig2

        fig2 = last_fig(
            color_code, first, temp, pos, pkeys, centrex, centrey, color_dict, labelsx=False
        )
        # plt.axis('off')
        # fig1.tight_layout()
        col1, col2 = st.beta_columns(2)

        col1.pyplot(fig1, use_column_width=True)
        col2.pyplot(fig2, use_column_width=True)
        fig2 = last_fig(
            color_code, first, temp, pos, pkeys, centrex, centrey, color_dict, labelsx=True
        )

        st.pyplot(fig2, use_column_width=True)

    # try:
    #    st.pyplot(fig1, use_column_width=True)
    #    st.pyplot(fig2, use_column_width=True)

    # except:

    #    fig2.savefig("img2.png")
    #    import matplotlib.image as mpimg
    #    img1 = mpimg.imread('img1.png')
    #    img2 = mpimg.imread('img2.png')

    #    ax1.imshow(img1)
    #    ax1.axis('off')
    #    ax2.imshow(img2)
    #    ax2.axis('off')


def list_centrality(first):
    H = first.to_undirected()
    st.markdown("## Betweeness Centrality:")
    st.markdown("Top to bottom node id from most central to least:")
    centrality = nx.betweenness_centrality(H, endpoints=True)

    # centrality = nx.betweenness_centrality(H)#, endpoints=True)
    df = pd.DataFrame([centrality])
    df = df.T
    df.sort_values(0, axis=0, ascending=False, inplace=True)
    df.rename(columns={0: "centrality value"}, inplace=True)

    bc = df
    st.markdown("### Most Connected:")
    st.write(bc.head())
    # st.text("...")
    # st.markdown("### Least Connected:")
    # st.write(bc.tail())

    st.markdown("## In degree Centrality:")  # " (percieved listeners/high authority)")
    st.markdown("Top to bottom node id from most central to least:")

    centrality = nx.in_degree_centrality(first)
    df = pd.DataFrame([centrality])
    df = df.T
    df.sort_values(0, axis=0, ascending=False, inplace=True)
    df.rename(columns={0: "centrality value"}, inplace=True)
    st.markdown("### Biggest targets (group acknowledged and verified communication):")

    st.write(df.head())
    # st.text("...")
    # st.markdown("### Least Listening:")

    # st.write(df.tail())

    st.markdown(
        "## Out-degree Centrality"
    )  # , read from top to bottom from most central to least:"
    # )

    centrality = nx.out_degree_centrality(first)
    df = pd.DataFrame([centrality])
    df = df.T
    df.sort_values(0, axis=0, ascending=False, inplace=True)
    df.rename(columns={0: "centrality value"}, inplace=True)
    st.markdown("### Biggest sources (individually acknowledged communication):")

    st.write(df.head())
    # st.text("...")
    # st.markdown("### Least Talkative:")

    # st.write(df.tail())

    # bc = df
    # st.table(df)
    return bc

    # Compute the in-degree centrality for nodes.
    # st.markdown("Out-degree Centrality:")
    # st.markdown("Top to bottom node id from most central to least:")

    # Compute the out-degree centrality for nodes.
    # st.markdown("Betweeness Centrality:")
    # centrality = nx.betweenness_centrality(H, endpoints=True)
    # df = pd.DataFrame([centrality])
    # df = df.T
    # df.sort_values(0, axis=0, ascending=False, inplace=True)
    # st.table(df)
    # edge_thickness = {k: v * 200000 for k, v in centrality.items()}


# @st.cache

# Or, if you know what you're doing, use
import fileinput
import sys

def replaceAll(file,searchExp,replaceExp):
    for line in fileinput.input(file, inplace=1):
        if searchExp in line:
            line = line.replace(searchExp,replaceExp)
        sys.stdout.write(line)


def physics(first, adj_mat_dicts, color_code, color_code_0, color_dict):

    #my_expander = st.expander("physical parameters")

    #phys_ = my_expander.radio(
    #    "Would you like to change physical parameters?", ("No", "Yes")
    #)

    # my_expander = st.sidebar.expander("Explanation of Threshold")
    my_expander2 = st.expander("Explanation")
    #physics_layouts = st.expander("Layouts")
    #physics_layouts.radio(
    #    label="layout_options", options=("force_atlas_2based", "hierarchical")
    #)
    my_expander2.markdown(
        """
    The specific interactive visualization libraries are: pyvis, which calls the javascript library: vis-network
    https://pyvis.readthedocs.io/en/latest/documentation.html
    https://github.com/visjs/vis-network
    Barnes Hut is a quadtree based gravity model.
    The barnesHut physics model (which is enabled by default) is based on an inverted gravity model. By increasing the mass of a node, you increase it’s repulsion. Values lower than 1 are not recommended.
    Other visualizations use a Fruchterman force directed layout algorithm.

    Almost all physical simulations use different random initial conditions.


    [LC Freeman A set of Measures of Betweeness (1977)](https://www.jstor.org/stable/pdf/3033543.pdf?casa_token=TzgYRJHfiYwAAAAA:r_8UKsxHRT7GRzoZ1OXwhJpzBbalTBYbG53me2fyMgZOvHnS9XM5TGB5yusfk5mCzQqXz4exAEFUcKXZ8I5ciIlU2dGpADzDfMu4Zm0rdA65G_ZzzJGo)
    [Analyzing the Structure of
    the Centrality-Productivity Literature
    Created Between 1948 and 1979](https://journals.sagepub.com/doi/pdf/10.1177/107554709001100405?casa_token=49LZA0RLipUAAAAA:nP4ZKyjVjgiuskFOE1540eeixMGwt0mW8-2VNCzfdV0IoRYFWSsrQLXTZAVWulawQqJ9A4XcND--Sw)\n
    Exploring network structure, dynamics, and function using NetworkX
    A Hagberg, P Swart, DS Chult - 2008 - osti.gov
    … NetworkX is a Python package for exploration and analysis of networks and network algorithms …
    NetworkX Python libraries to extend the avail- able functionality with interfaces to well-tested
    numerical and statis- tical libraries written in C. C++ and FORTRAN …
      Cited by 3606 Related articles


    The basic force directed layout (used in other visualizations not this one)
    Qoute from wikipedia:
    'Force-directed graph drawing algorithms assign forces among the set of edges and the set of nodes of a graph drawing. Typically, spring-like attractive forces based on Hooke's law are used to attract pairs of endpoints of the graph's edges towards each other, while simultaneously repulsive forces like those of electrically charged particles based on Coulomb's law are used to separate all pairs of nodes. In equilibrium states for this system of forces, the edges tend to have uniform length (because of the spring forces), and nodes that are not connected by an edge tend to be drawn further apart (because of the electrical repulsion). Edge attraction and vertex repulsion forces may be defined using functions that are not based on the physical behavior of springs and particles; for instance, some force-directed systems use springs whose attractive force is logarithmic rather than linear.'
    \n
    https://en.wikipedia.org/wiki/Force-directed_graph_drawing \n
    What this means is conflicting forces of attraction, and repulsion determine node position.
    Possesing a high centrality value does not necessarily mean occupying a central position on the graph.
    Also nodes can have a high betweeness centrality due to contributions from either inward directed projections, outward facing projections or both.\n

    Fruchterman, Thomas M. J.; Reingold, Edward M. (1991), "Graph Drawing by Force-Directed Placement", Software – Practice & Experience, Wiley, 21 (11): 1129–1164, doi:10.1002/spe.4380211102.

    """
    )
    #my_expander = st.expander("Mouse over node info?")

    #mo_ = my_expander.radio("Toggle Mouse overs?", ("Yes", "No"))
    #if mo_ == "Yes":
    #        mo = True
    #else:
    #    mo = False
    mo = True

    #my_expander = st.expander("Directed Visualization (include arrow directions)?")

    #dir_ = my_expander.radio("Toggle Directed Visualization?", ("No","Yes"))
    #if dir_ == "Yes":
    #    dir = True
    #else:
    #    dir = False

    dir = True
    pos = nx.get_node_attributes(first, "pos")

    nt = Network(
        notebook=True,
        directed=True,
        height="500px",
        width="100%",
        font_color="black",  # , bgcolor='#222222'
    )  # bgcolor='#222222',
    if dir:
        nt = Network("800px", "800px", directed=True,font_color="black")  # ,layout=physics_layouts)
    else:
        nt = Network("800px", "800px", directed=False,font_color="black")  # ,layout=physics_layouts)

    #nt.set_edge_smooth('continuous')

    #nt.barnes_hut()
    for node in first.nodes:
        nt.add_node(node,label=node) # node id = 1 and label = Node 1

    for e in first.edges:
        src = e[0]
        dst = e[1]
        src = str(src)
        dst = str(dst)

        ee = first.get_edge_data(e[0], e[1])
        #nt.add_edge(src, dst, arrowStrikethrough=True)

        nt.add_edge(src, dst, width=0.4*ee["weight"], arrowStrikethrough=True)
    nt.inherit_edge_colors(True)
    H = first.to_undirected()
    # centrality = nx.betweenness_centrality(H)#, k=10, endpoints=True)
    centrality = nx.betweenness_centrality(H, endpoints=True)

    edge_thickness = {k: v * 90000000 for k, v in centrality.items()}
    node_size = {k: v*320 for k, v in centrality.items()}


    neighbor_map = nt.get_adj_list()
    for node in nt.nodes:
        if dir:
            node["size"] = node_size[node["id"]] #* 10025
        node["borderWidth"] = 2

    # add neighbor data to node hover data
    for node in nt.nodes:
        if mo:
            if "title" not in node.keys():
                if node["id"] in color_code_0.keys():
                    node["title"] = (
                        "<br> This node is:"
                        + str(node["id"])
                        + "<br> it's membership is "
                        + str(color_code_0[node["id"]])
                        + "<br> It's neighbors are:<br>"
                        + "<br>".join(neighbor_map[node["id"]])
                    )
                else:
                    node["title"] = (
                        "<br> This node is:"
                        + str(node["id"])
                        + "<br> it's membership is "
                        + str("unknown")
                        + " It's neighbors are:<br>"
                        + "<br>".join(neighbor_map[node["id"]])
                    )
        if node["id"] in node_size.keys():
            node["size"] = node_size[node["id"]]
        node["label"] = str(node["id"])
        if not dir:
            node["value"] = 10.0*len(neighbor_map[node["id"]])
        if node["id"] in color_code.keys():
            node["color"] = color_code[node["id"]]
    nt.barnes_hut()
    #if phys_ == "Yes":
    #    nt.show_buttons(filter_=["physics"])
    def display():
        nt.save_graph("test1.html")

        #replaceAll('/test1.html','updateInterval": 50"','updateInterval": 50000"')
        HtmlFile = open("test1.html", "r", encoding="utf-8")
        source_code = HtmlFile.read()
        components.html(source_code, height=800, width=800)  # ,use_column_width=True)

    if dir:
        #st.markdown("""
        #This visualization has bigger node sizes because it uses betweeness centrality to determine node size.
        #The other visualization determines node size based on just a count on the number of direct neighbours
        #""")
        #replaceAll('/saved_html2.html','updateInterval": 50"','updateInterval": 50000"')

        nt.save_graph("saved_html2.html")
        HtmlFile = open("saved_html2.html", "r", encoding="utf-8")
        source_code = HtmlFile.read()
        components.html(source_code, height=800, width=800)  # ,use_column_width=True)

    else:
        display()

    #except:
    #    pass
    # @st.cache(suppress_st_warning=True)# to suppress the warning.


    #    network = drawGraph();
    #    localStorage.setItem("localnet",network)

    #if phys_ == "Yes":
    #    from PIL import Image

    #    st.markdown(
    #        "Some parameter sets can prevent static equilibrium states. For example:"
    #    )
        # nt.show_buttons(filter_=["physics"])
    #    st.image(Image.open("rescreen_shot_just_params.png"))


def dont():
    chord = hv.Chord(links)  # .select(value=(5, None))
    # node_color = [color_code[n] for n in H]
    # st.text(links['color'])
    chord.opts(
        opts.Chord(
            cmap="Category20",
            width=250,
            height=250,
            edge_cmap="Category20",
            edge_color=dim("source").str(),
            labels="name",
            node_color=dim("index").str(),
        )
    )


from scipy.spatial import Delaunay, ConvexHull
from pyveplot import Hiveplot, Axis, Node
import networkx as nx
import random
import base64
import textwrap


def render_svg_small(svg):
    """Renders the given svg string."""
    b64 = base64.b64encode(svg.encode("utf-8")).decode("utf-8")
    html = r'<img src="data:image/svg+xml;base64,%s" width = 900/>' % b64
    st.write(html, unsafe_allow_html=True, use_column_width=True)
    return None


#        hub_sort(first,color_code_0,reverse)
def render_svg(svg):
    """Renders the given svg string."""
    b64 = base64.b64encode(svg.encode("utf-8")).decode("utf-8")
    html = r'<img src="data:image/svg+xml;base64,%s" width = 900/>' % b64
    st.write(html, unsafe_allow_html=True, use_column_width=True)
    return None


def agraph_(first):
    from streamlit_agraph import agraph, Node, Edge, Config

    config = Config(
        height=500,
        width=700,
        nodeHighlightBehavior=True,
        highlightColor="#F7A7A6",
        directed=True,
        collapsible=True,
    )
    # st.text(dir(agraph))
    # agraph(list(first.nodes), (first.edges), config)


def hub_sort(first, color_code_1, reverse):
    with open("1ba_hiveplot.svg", "r") as f:
        lines = f.readlines()
        f.close()
    line_string = "".join(lines)

    render_svg_small(line_string)
    line_string = None
    lines = None
    del line_string
    del lines
    return None

    """
    c = ['#e41a1c', '#377eb8', '#4daf4a',
         '#984ea3', '#ff7f00', '#ffff33',
         '#a65628', '#f781bf', '#999999',]

    # create hiveplot object
    import svgwrite
    dwg = svgwrite.Drawing()
    del dwg
    h = None
    h = Hiveplot()
    h.__init__()
    #fig = plt.figure()
    h.axes = None
    # create three axes, spaced at 120 degrees from each other
    h.axes = [Axis(start=20, angle=0,
                   stroke='black', stroke_width=2.1),
              Axis(start=20, angle=90,
                   stroke='black', stroke_width=2.1),
              Axis(start=20, angle=90 + 90,
                   stroke='black', stroke_width=2.1),
              Axis(start=20, angle=90 + 90 + 90,
                   stroke='black', stroke_width=2.1)

              ]

    # create a random Barabasi-Albert network
    g = first

    # sort nodes by degree
    k = list(nx.degree(g))
    k.sort(key=lambda tup: tup[1])

    maxd = np.max([i[1] for i in k])

    # categorize them as high, medium and low degree
    hi_deg = [v[0] for v in k if v[1] > 3*maxd/4]
    md_deg = [v[0] for v in k if v[1] > maxd/4 and v[1] <= 2*maxd/4]
    md_deg2 = [v[0] for v in k if v[1] > 2*maxd/4 and v[1] <= 3*maxd/4]
    lo_deg = [v[0] for v in k if v[1] <= maxd/4]

    # place these nodes into our three axes
    for axis, nodes in zip(h.axes,
                           [hi_deg, md_deg,md_deg2, lo_deg]):
        #random.choice(c)
        for v in nodes:
            circle_color = color_code_1[v]
            # create node object
            node = Node(radius=22.5*g.degree(v),
                        label="%s" % (v))
            # add it to axis
            axis.add_node(v, node)
            # once it has x, y coordinates, add a circle
            node.add_circle(fill=circle_color, stroke=circle_color,
                            stroke_width=0.1, fill_opacity=0.65)
            if axis.angle < 180:
                orientation = -1 #1#-1
                scale = 8.5
            else:
                orientation = 1
                scale = 1.5

                if axis.angle <5 or axis.angle>355:
                    orientation = 1
                    scale = 7.5
            # also add a label
            node.add_label("{0}".format(v),
                           angle=axis.angle + 90 * orientation,
                           scale=scale)
            #st.text("node {0}".format(v))

    # iterate through axes, from left to right
    for n in range(-1, len(h.axes) - 1):

        curve_color = 'black'#random.choice(c)
        # draw curves between nodes connected by edges in network
        h.connect_axes(h.axes[n],
                       h.axes[n+1],
                       g.edges,
                       stroke_width=4.5,
                       stroke=curve_color)

    # save output
    import svgwrite
    dwg = svgwrite.Drawing()
    del dwg

    import os
    os.system('rm ba_hiveplot.svg')
    h.save(str(a)+'ba_hiveplot.svg')
    del h
    h = None
    #h = None
    h = Hiveplot()
    h.__init__()
    fig = plt.figure()
    """

    # line_string = ''
    # os.system('rm ba_hiveplot.svg')
    # from streamlit import caching

    # caching.clear_cache()


# a = 0


def hive_two(first, color_code, color_code_0, reverse):

    with open("0ba1_hiveplot.svg", "r") as f:
        lines = f.readlines()
        f.close()
    line_string = "".join(lines)

    render_svg_small(line_string)
    line_string = None
    del line_string
    # os.system("rm ba1_hiveplot.svg")
    # from streamlit import caching

    # caching.clear_cache()
    return None


def no_thanks():
    from hiveplotlib import Axis, Node, HivePlot

    # convert `networkx` edges and nodes into `hiveplotlib`-ready structures
    G = first
    encoded = {v: k for k, v in enumerate(first.nodes())}
    reverse = {v: k for k, v in encoded.items()}

    G = nx.relabel_nodes(G, encoded, copy=True)
    edges = np.array(G.edges)

    # pull out degree information from nodes for later use
    node_ids, degrees = np.unique(edges, return_counts=True)

    # nodes = np.array(G.nodes)
    nodes = []

    IRG1_indices = []
    IRG2_indices = []
    IRG3_indices = []
    DCMT_ind = []  # ,Un_ind
    # st.text(len(color_code_0))
    for i, (node_id, degree) in enumerate(zip(node_ids, degrees)):
        if not reverse[node_id] in color_code_0.keys():
            color_code_0[reverse[node_id]] = hc[reverse[node_id]]
            reverse[node_id] = hc[reverse[node_id]]

        temp_node = Node(unique_id=node_id, data=G.nodes.data()[node_id])
        nodes.append(temp_node)
    for i, (node_id, degree) in enumerate(zip(node_ids, degrees)):
        # store the index number as a way to align the nodes on axes
        G.nodes.data()[node_id]["loc"] = node_id
        # also store the degree of each node as another way to align nodes on axes
        G.nodes.data()[node_id]["degree"] = degree
        # G.nodes.data()[node_id]['club'] =

        if reverse[node_id] in color_code_0.keys():
            # = color_code_0[reverse[node_id]]
            if color_code_0[reverse[node_id]] == "IRG 1":
                IRG1_indices.append(i)
                G.nodes.data()[node_id]["IRG 1"] = 1
                G.nodes.data()[node_id]["club"] = 1
            if color_code_0[reverse[node_id]] == "IRG 2":
                IRG2_indices.append(i)
                G.nodes.data()[node_id]["IRG 2"] = 2
                G.nodes.data()[node_id]["club"] = 2

            if color_code_0[reverse[node_id]] == "IRG 3":
                IRG3_indices.append(i)
                G.nodes.data()[node_id]["IRG 3"] = 3
                G.nodes.data()[node_id]["club"] = 3

            # st.text(IRG3_indices)
            if color_code_0[reverse[node_id]] == "DCMT":
                DCMT_ind.append(i)
                G.nodes.data()[node_id]["DCMT"] = 4
                G.nodes.data()[node_id]["club"] = 4

    temp = list(set(color_code_0.values()))
    hp = hive_plot_n_axes(
        node_list=nodes,
        edges=edges,
        axes_assignments=[
            IRG1_indices,
            IRG2_indices,
            IRG3_indices,
            DCMT_ind,
        ],
        sorting_variables=["club", "club", "club", "club"],
        axes_names=temp,
        vmins=[0, 0, 0, 0],
        vmaxes=[2, 2, 2, 2],
        orient_angle=30,
    )

    # change the line kwargs for edges in plot
    hp.add_edge_kwargs(
        axis_id_1=temp[0], axis_id_2=temp[1], c=f"C0", lw=1.5, alpha=0.5, zorder=10
    )
    hp.add_edge_kwargs(
        axis_id_1=temp[1], axis_id_2=temp[2], c=f"C2", lw=1.5, alpha=0.5, zorder=10
    )
    # hp.add_edge_kwargs(
    ##    axis_id_1=temp[0], axis_id_3=temp[2], c=f"C1", lw=1.5, alpha=0.5, zorder=10
    # )

    # st.text(temp[2])
    hp.place_nodes_on_axis(
        axis_id=temp[0],
        unique_ids=[nodes[i].data["loc"] for i in IRG1_indices],
        sorting_feature_to_use="loc",
        vmin=0,
        vmax=33,
    )
    hp.place_nodes_on_axis(
        axis_id=temp[1],
        unique_ids=[nodes[i].data["loc"] for i in IRG2_indices],
        sorting_feature_to_use="loc",
        vmin=0,
        vmax=33,
    )
    hp.place_nodes_on_axis(
        axis_id=temp[2],
        unique_ids=[nodes[i].data["loc"] for i in IRG3_indices],
        sorting_feature_to_use="loc",
        vmin=0,
        vmax=33,
    )
    hp.place_nodes_on_axis(
        axis_id=temp[3],
        unique_ids=[nodes[i].data["loc"] for i in DCMT_ind],
        sorting_feature_to_use="loc",
        vmin=0,
        vmax=33,
    )
    # hp.place_nodes_on_axis(
    #    axis_id=temp[3],
    #    unique_ids=[nodes[i].data["loc"] for i in Un_ind],
    #    sorting_feature_to_use="loc",
    #    vmin=0,
    #    vmax=33,
    # )

    hp.connect_axes(edges=edges, axis_id_1=temp[0], axis_id_2=temp[1], c="C1")
    hp.connect_axes(edges=edges, axis_id_1=temp[1], axis_id_2=temp[2], c="C2")
    hp.connect_axes(edges=edges, axis_id_1=temp[0], axis_id_2=temp[2], c="C2")
    hp.connect_axes(edges=edges, axis_id_1=temp[2], axis_id_2=temp[3], c="C3")
    hp.connect_axes(edges=edges, axis_id_1=temp[3], axis_id_2=temp[1], c="C1")
    hp.connect_axes(edges=edges, axis_id_1=temp[3], axis_id_2=temp[0], c="C0")
    # hp.connect_axes(edges=edges, axis_id_1=temp[4], axis_id_2=temp[0], c="C7")
    # hp.connect_axes(edges=edges, axis_id_1=temp[4], axis_id_2=temp[1], c="C8")
    # hp.connect_axes(edges=edges, axis_id_1=temp[4], axis_id_2=temp[2], c="C9")
    # hp.connect_axes(edges=edges, axis_id_1=temp[4], axis_id_2=temp[3], c="C10")

    fig, ax = hive_plot_viz_mpl(hive_plot=hp)
    st.pyplot(fig)


import networkx as nx
import numpy as np
from scipy import integrate


def disparity_filter(G, weight="weight"):
    """Compute significance scores (alpha) for weighted edges in G as defined in Serrano et al. 2009
    Args
        G: Weighted NetworkX graph
    Returns
        Weighted graph with a significance score (alpha) assigned to each edge
    References
        M. A. Serrano et al. (2009) Extracting the Multiscale backbone of complex weighted networks. PNAS, 106:16, pp. 6483-6488.
    """

    if nx.is_directed(G):  # directed case
        N = nx.DiGraph()
        for u in G:

            k_out = G.out_degree(u)
            k_in = G.in_degree(u)

            if k_out > 1:
                sum_w_out = sum(np.absolute(G[u][v][weight]) for v in G.successors(u))
                for v in G.successors(u):
                    w = G[u][v][weight]
                    p_ij_out = float(np.absolute(w)) / sum_w_out
                    alpha_ij_out = (
                        1
                        - (k_out - 1)
                        * integrate.quad(lambda x: (1 - x) ** (k_out - 2), 0, p_ij_out)[
                            0
                        ]
                    )
                    N.add_edge(u, v, weight=w, alpha_out=float("%.4f" % alpha_ij_out))

            elif k_out == 1 and G.in_degree(G.successors(u)) == 1:
                # we need to keep the connection as it is the only way to maintain the connectivity of the network
                v = G.successors(u)[0]
                w = G[u][v][weight]
                N.add_edge(u, v, weight=w, alpha_out=0.0, alpha_in=0.0)
                # there is no need to do the same for the k_in, since the link is built already from the tail

            if k_in > 1:
                sum_w_in = sum(np.absolute(G[v][u][weight]) for v in G.predecessors(u))
                for v in G.predecessors(u):
                    w = G[v][u][weight]
                    p_ij_in = float(np.absolute(w)) / sum_w_in
                    alpha_ij_in = (
                        1
                        - (k_in - 1)
                        * integrate.quad(lambda x: (1 - x) ** (k_in - 2), 0, p_ij_in)[0]
                    )
                    N.add_edge(v, u, weight=w, alpha_in=float("%.4f" % alpha_ij_in))
        return N

    else:  # undirected case
        B = nx.Graph()
        for u in G:
            k = len(G[u])
            if k > 1:
                sum_w = sum(np.absolute(G[u][v][weight]) for v in G[u])
                for v in G[u]:
                    w = G[u][v][weight]
                    p_ij = float(np.absolute(w)) / sum_w
                    alpha_ij = (
                        1
                        - (k - 1)
                        * integrate.quad(lambda x: (1 - x) ** (k - 2), 0, p_ij)[0]
                    )
                    B.add_edge(u, v, weight=w, alpha=float("%.4f" % alpha_ij))
        return B


def disparity_filter_alpha_cut(G, weight="weight", alpha_t=0.4, cut_mode="or"):
    """Performs a cut of the graph previously filtered through the disparity_filter function.

    Args
    ----
    G: Weighted NetworkX graph

    weight: string (default='weight')
        Key for edge data used as the edge weight w_ij.

    alpha_t: double (default='0.4')
        The threshold for the alpha parameter that is used to select the surviving edges.
        It has to be a number between 0 and 1.

    cut_mode: string (default='or')
        Possible strings: 'or', 'and'.
        It works only for directed graphs. It represents the logic operation to filter out edges
        that do not pass the threshold value, combining the alpha_in and alpha_out attributes
        resulting from the disparity_filter function.


    Returns
    -------
    B: Weighted NetworkX graph
        The resulting graph contains only edges that survived from the filtering with the alpha_t threshold

    References
    ---------
    .. M. A. Serrano et al. (2009) Extracting the Multiscale backbone of complex weighted networks. PNAS, 106:16, pp. 6483-6488.
    """

    if nx.is_directed(G):  # Directed case:
        B = nx.DiGraph()
        for u, v, w in G.edges(data=True):
            try:
                alpha_in = w["alpha_in"]
            except KeyError:  # there is no alpha_in, so we assign 1. It will never pass the cut
                alpha_in = 1
            try:
                alpha_out = w["alpha_out"]
            except KeyError:  # there is no alpha_out, so we assign 1. It will never pass the cut
                alpha_out = 1

            if cut_mode == "or":
                if alpha_in < alpha_t or alpha_out < alpha_t:
                    B.add_edge(u, v, weight=w[weight])
            elif cut_mode == "and":
                if alpha_in < alpha_t and alpha_out < alpha_t:
                    B.add_edge(u, v, weight=w[weight])
        return B

    else:
        B = nx.Graph()  # Undirected case:
        for u, v, w in G.edges(data=True):

            try:
                alpha = w["alpha"]
            except KeyError:  # there is no alpha, so we assign 1. It will never pass the cut
                alpha = 1

            if alpha < alpha_t:
                B.add_edge(u, v, weight=w[weight])
        return B


def main():

    #st.write(tkinter.TkVersion)
    #root = tkinter.Tk()

    # full range.
    #"Physics",
    # "3D",
    # "Population",
    #"Interactive Population",
    #"Visualize Centrality",
    #"Hive",
    #"Community Mixing",
    #"Basic",
    #"Spreadsheet",
    #"Bundle",
    #"AdjacencyMatrix",
    #"Chord",
    #"View Source Code",

    st.sidebar.title("Odor To Action: Collaboration Survey Data")

    genre = st.sidebar.radio(
        "Choose Graph Layout/Option:",
        (
            "Physics",
            "Interactive Population",
            "Visualize Centrality",
            "Hive",
            "Community Mixing",
            "Basic",
            "Spreadsheet",
            "Bundle",
            "AdjacencyMatrix",
            "Chord",
            "3D",
            "View Source Code",
        ),
    )

    my_expander = st.sidebar.expander("Explanation of Threshold")

    my_expander.markdown(
        """
		Problem most people politely answer that they talk to someone a little bit, \
		a bias if which is not corrected \n for hyperconnects everyone to everyone \
		else in a meaningless way. The solution \n to this problem involves thresholding, \
		setting a minimum meaningful level of \n communication collaboration, \
		The higher the threshold the more you \n reduce connections"""
    )
    #my_expander = st.expander("Toggle Transpose collaboration sources/targets")
    #transpose = my_expander.radio("source/target", (False, True))
    transpose = False
    my_expander = st.expander("Set threshold")
    threshold = my_expander.slider("Select a threshold value", 0.0, 8.0, 5.0, 1.0)
    (
        df2,
        names,
        ratercodes,
        legend,
        color_code,
        color_dict,
        color_code_0,
        sheet,
        popg,
        hc,
    ) = get_frame(transpose, threshold)

    fig = plt.figure()
    for k, v in color_dict.items():
        plt.scatter([], [], c=v, label=k, s=350)
    plt.legend(frameon=False, prop={"size": 35})
    fig.tight_layout()
    #plt.axis("off")
    st.sidebar.markdown("Color coding of most plots")
    # my_expander.markdown(
    #    """ Excepting for chord and hive, which are time consuming to code"""
    # )
    st.sidebar.pyplot(fig)
    inboth = set(names) & set(ratercodes)
    notinboth = set(names) - set(ratercodes)
    allcodes = set(names) or set(ratercodes)
    first = nx.DiGraph()

    for i, row in enumerate(allcodes):
        if i != 0:
            if row[0] != 1 and row[0] != 0:
                first.add_node(row[0], name=row)  # ,size=20)
    adj_mat_dicts = []
    conns = {}
    cc = copy.copy(color_code_0)
    for i, idx in enumerate(df2.index):
        for j, col in enumerate(df2.columns):
            if col not in cc.keys():
                cc[col] = hc[col]
            if idx not in color_code_0.keys():
                cc[col] = hc[col]

    for i, idx in enumerate(df2.index):
        for j, col in enumerate(df2.columns):
            weight = float(df2.iloc[i, j])
            if idx != col:
                if float(weight) > threshold:
                    adj_mat_dicts.append({"src": idx, "tgt": col, "weight": weight})
                    first.add_edge(idx, col, weight=weight)

            if not popg.has_edge(cc[idx], cc[col]):
                popg.add_edge(cc[idx], cc[col], weight=weight)
            else:
                e = popg.get_edge_data(cc[idx], cc[col])
                weight = weight + e["weight"]
                popg.add_edge(cc[idx], cc[col], weight=weight)

    first.remove_nodes_from(list(nx.isolates(first)))
    adj_mat = pd.DataFrame(adj_mat_dicts)
    try:
        encoded = {v: k for k, v in enumerate(first.nodes())}
    except:
        encoded = {v: k for k, v in enumerate(adj_mat.columns)}
    # adj_mat = adj_mat[adj_mat["weight"] != 0]

    link = dict(
        source=[encoded[i] for i in list(adj_mat["src"].values)],
        target=[encoded[i] for i in list(adj_mat["tgt"].values)],
        value=[i * 3 for i in list(adj_mat["weight"].values)],
    )
    adj_mat2 = pd.DataFrame(link)
    adj_mat3 = adj_mat[adj_mat["weight"] != 0]

    encoded = {v: k for k, v in enumerate(first.nodes())}
    reverse = {v: k for k, v in encoded.items()}
    G = nx.relabel_nodes(first, encoded, copy=True)
    edges = np.array(G.edges)

    node_ids, degrees = np.unique(edges, return_counts=True)

    for i, (node_id, degree) in enumerate(zip(node_ids, degrees)):
        if not reverse[node_id] in color_code_0.keys():
            color_code_0[reverse[node_id]] = hc[reverse[node_id]]
            reverse[node_id] = hc[reverse[node_id]]
    if genre == "View Source Code":
        st.markdown("""https://github.com/russelljjarvis/odor2action""")

        st.markdown(
            """[mostly in this file](https://github.com/russelljjarvis/odor2action/blob/master/app.py)"""
        )
    if genre == "Visualize Centrality":

        my_expander = st.expander("Explanation of Centrality Hive")

        my_expander.markdown(
            """Using pythons networkx module Nodes are layed out from ascending to descending contributions of centrality. This plot depicts betweeness centrality from densely inter-connected (hub) to sparsely inter-connected leaf.
            Hive visualizations are designed to show between group connectivity. Nodes on the same axis have implied connectivity through the axis, these connections are not shown to remove clutter."""
        )
        # Unlike other hive implementations nodes are shown that don't connect outside of their assigned group (eg IRG-1), but they are visualized with no prominent interconnection.
        # Connection within their class (eg IRG-1 is implied by being positioned on a vertical or horizontal axis, the axis connects all nodes.
        # """

        #try:
        #    import os

            #if transpose:
            #    os.system("python make_serial_transpose.py")

            #else:
            #    os.system("python make_serial_plots1.py")
        #except:
        #    pass
        hub_sort(first, color_code, reverse)
        list_centrality(first)
    if genre == "Spreadsheet":
        st.markdown("Processed anonymized network data that is visualized")
        st.markdown(get_table_download_link_csv(df2), unsafe_allow_html=True)
        st.markdown("Anonymized raw survey data")
        st.markdown(get_table_download_link_csv(sheet), unsafe_allow_html=True)
        my_expander = st.expander("Numeric mapping of survery question answers")
        my_expander.write(legend)
        my_expander = st.expander("Collapsed/Expand Numeric Spread sheet")
        my_expander.table(df2)

        my_expander = st.expander("Collapsed/Expand Raw Spread sheet")
        my_expander.markdown(
            """Row elements represent, outward facing projections, how often a person recognizes contact with others."""
        )
        my_expander.markdown(
            """Column elements represent, inward facing projections, how often other people recognize contact with this person"""
        )

        my_expander.table(sheet)
        my_expander = st.expander("Verify Person By Code")
        user_input = my_expander.text_input("enter anonymos code", "02P1")
        try:
            my_expander.markdown("The column is how others rated them")

            my_expander.write(df2[user_input])
            my_expander.markdown(
                "If following row slice has no value, this user has not participated in survey, but people have answered questions about them"
            )

            my_expander.write(df2.loc[df2.index.isin([user_input])])

        except:
            my_expander.warning(
                "This user has not participated in survey, but people have answered questions about them"
            )
            my_expander.warning("Try toggling the transpose")
            my_expander.write(df2[user_input])
            my_expander.write(df2.loc[df2.index.isin([user_input])])

            # my_expander.write(df2.loc[:,user_input])
    if genre == "Community Mixing":
        my_expander = st.expander("Explanation of Community Partitions")
        my_expander.markdown(
            """Communities in the graph on the left are not IRG 1-3, but instead communities found by blind network analysis. It's appropritate to use a different color code for the four inferred communities. \
        For contrast in the graph on the right, machine driven community detection clusters persist, but now nodes (dots) are color coded IRG-1-3 \n \
        This suggests that the formal memberships eg. \"IRG 1\" does not determine the machine generated communities. In otherwords spontaneuosly emerging community groups may be significantly different to formal group assignments.
        The stochastic community detection algorithm uses a differently seeded random number generator every time so the graph appears differently each time the function is called.
        The algorithm is called Louvain community detection. The Louvain Community algorithm detects 5 communities, but only 2 communities with membership >=3. A grey filled convex hull is drawn around each of the two larger communities.
        """
        )

        community(first, color_code, color_dict)
    '''
    if genre == "3D":
        st.markdown("""This visualization is a work in progress, the results are not yet fully reliable""")
        my_expander = st.expander("Explanation of different Node sizes for 3D")
        my_expander.markdown(
            """Note this visualization uses a different library (I-Graph, however normally networkx is used) to determine the betweeness centrality
        with the interesting consequence that a DCMT node is now second most central"""
        )
        # g = first

        # links = copy.copy(adj_mat)
        # links.rename(
        #    columns={"weight": "value", "src": "source", "tgt": "target"}, inplace=True
        # )
        # links = links[links["value"] != 0]
        # Edges = [
        #    (encoded[src], encoded[tgt])
        #    for src, tgt in zip(links["source"], links["target"])
        # ]
        # G = ig.Graph(Edges, directed=True)
        #G = ig.Graph.from_networkx(first)  # , directed=True)
        # local_reverse = {k:v for k,v in zip(G.vs(),first.nodes)}
        # spring_3D = nx.spring_layout(G, dim = 3, k = 0.5) # k regulates the distance between nodes
        edges = first.edges()

        # ## update to 3d dimension
        spring_3D = nx.spring_layout(
            first, dim=3, scale=2.5, k=0.0005
        )  # k regulates the distance between nodes
        # weights = [G[u][v]['weight'] for u,v in edges]
        # nx.draw(G, with_labels=True, node_color='skyblue', font_weight='bold',  width=weights, pos=pos)

        # we need to seperate the X,Y,Z coordinates for Plotly
        # NOTE: spring_3D is a dictionary where the keys are 1,...,6
        x_nodes = [
            spring_3D[key][0] for key in spring_3D.keys()
        ]  # x-coordinates of nodes
        y_nodes = [spring_3D[key][1] for key in spring_3D.keys()]  # y-coordinates
        z_nodes = [spring_3D[key][2] for key in spring_3D.keys()]  # z-coordinates

        # we need to create lists that contain the starting and ending coordinates of each edge.
        x_edges = []
        y_edges = []
        z_edges = []

        # create lists holding midpoints that we will use to anchor text
        xtp = []
        ytp = []
        ztp = []

        # need to fill these with all of the coordinates
        for edge in edges:
            # format: [beginning,ending,None]
            x_coords = [spring_3D[edge[0]][0], spring_3D[edge[1]][0], None]
            x_edges += x_coords
            xtp.append(0.5 * (spring_3D[edge[0]][0] + spring_3D[edge[1]][0]))

            y_coords = [spring_3D[edge[0]][1], spring_3D[edge[1]][1], None]
            y_edges += y_coords
            ytp.append(0.5 * (spring_3D[edge[0]][1] + spring_3D[edge[1]][1]))

            z_coords = [spring_3D[edge[0]][2], spring_3D[edge[1]][2], None]
            z_edges += z_coords
            ztp.append(0.5 * (spring_3D[edge[0]][2] + spring_3D[edge[1]][2]))
        labels = []
        group = []
        human_group = []
        # local_reverse = {i:e.source for e,i in zip(G.es,first.edges)}

        # for node in first.nodes:#G.vs():#links["source"]:
        for node in first.nodes:
            labels.append(str(node) + str(" ") + str(color_code_0[node]))
            group.append(color_code[node])
            human_group.append(color_code_0[node])

        # etext = [f'weight={w}' for w in edge_weights]

        trace_weights = go.Scatter3d(
            x=xtp,
            y=ytp,
            z=ztp,
            mode="markers",
            marker=dict(color="rgb(125,125,125)", size=1),
        )  # set the same color as for the edge lines
        # )

        # create a trace for the edges
        trace_edges = go.Scatter3d(
            x=x_edges,
            y=y_edges,
            z=z_edges,
            mode="lines",
            line=dict(color="black", width=2),
            hoverinfo="none",
        )

        # create a trace for the nodes
        trace_nodes = go.Scatter3d(
            x=x_nodes,
            y=y_nodes,
            z=z_nodes,
            mode="markers",
            marker=dict(symbol="circle", size=10, color="skyblue"),
        )

        # Include the traces we want to plot and create a figure
        data = [trace_edges, trace_nodes, trace_weights]

        #layt = G.layout(
        #    "kk", dim=3
        #)  # plot network with the Kamada-Kawai layout algorithm
        #estimate = G.betweenness(directed=True)  # , cutoff=16)
        #H = first.to_undirected()

        # estimate = nx.betweenness_centrality(H)  # , k=10, endpoints=True)
        # estimate = list(estimate.values())
        ee = []
        for i in estimate:
            if i == 0:
                ee.append(30 * 0.45)
            else:
                ee.append(i * 0.45)
        estimate = ee
        # widths = []
        # for e in links["value"]:
        #    widths.append(1.85 * e)

        labels = []
        group = []
        human_group = []
        # local_reverse = {i:e.source for e,i in zip(G.es,first.edges)}

        # for node in first.nodes:#G.vs():#links["source"]:
        for node in first.nodes:
            labels.append(str(node) + str(" ") + str(color_code_0[node]))
            group.append(color_code[node])
            human_group.append(color_code_0[node])

        Xn = []
        Yn = []
        Zn = []
        N = len(first.nodes)
        for k in range(N):
            Xn += [layt[k][0] - 350]
            Yn += [layt[k][1]]
            Zn += [layt[k][2]]

        Xe = []
        Ye = []
        Ze = []
        group2 = []
        decoded = {v: k for k, v in encoded.items()}
        for e in G.es:
            Xe += [
                layt[e.source][0] - 350,
                layt[e.target][0] - 350,
                None,
            ]  # x-coordinates of edge ends
            Ye += [layt[e.source][1], layt[e.target][1], None]
            Ze += [layt[e.source][2], layt[e.target][2], None]
        # ,colorscale='Viridis'
        for e in G.es:
            group2.append(color_code[reverse[e.target]])

        trace1 = go.Scatter3d(
            x=Xe, y=Ye, z=Ze, mode="lines", line=dict(color="black", width=2.9)
        )  # ,text=labels,hoverinfo='text'))

        trace2 = go.Scatter3d(
            x=Xn,
            y=Yn,
            z=Zn,
            mode="markers",
            name="Researchers",
            marker=dict(
                symbol="circle",
                color=group,
                size=estimate,
                line=dict(color="rgb(50,50,50)", width=2),
            ),
            text=labels,
            hoverinfo="text",
        )

        axis = dict(
            showbackground=False,
            showline=False,
            zeroline=False,
            showgrid=False,
            showticklabels=False,
            title="",
        )

        layout = go.Layout(
            title="A 3D Visualization (can be rotated)",
            width=1200,
            height=1200,
            showlegend=False,
            scene=dict(
                xaxis=dict(axis),
                yaxis=dict(axis),
                zaxis=dict(axis),
            ),
        )

        data = [trace1, trace2]

        fig0 = go.Figure(data=data, layout=layout)
        st.write(fig0, use_column_width=True)
        # st.markdown("To ultimately be replaced with something in centre of page and networkx derived like below:")
        # fig1 = go.Figure(data=data, layout=layout)
        # st.write(fig1, use_column_width=True)
    '''
    if genre == "Physics":
        # if threshold == 5 and transpose == False:
        #    HtmlFile = open("test2.html", "r", encoding="utf-8")
        #    source_code = HtmlFile.read()
        #    components.html(
        #        source_code, height=750, width=750
        #    )  # ,use_column_width=True)

        # else:
        physics(first, adj_mat_dicts, color_code, color_code_0, color_dict)

    if genre == "Interactive Population":
        my_expander = st.expander("Explanation of population")
        my_expander.markdown(
            """Here node size does not reflect centrality or connectivity. Node size reflects number of participants in group, therefore DCMT is small because it consists of just two members. Likewise ribbon width is the total sum of weighted connections between groups."""
        )

        interactive_population(cc, popg, color_dict)

    if genre == "Population":
        my_expander = st.expander("Explanation of population")
        my_expander.markdown(
            """Here node size does not reflect centrality or connectivity. Node size reflects number of participants in group, therefore DCMT is small because it consists of just two members. Likewise ribbon width is the total sum of weighted connections between groups."""
        )

        population(cc, popg, color_dict)
    if genre == "Hive":

        my_expander = st.expander("Explanation of Hive")

        my_expander.markdown(
            """This predominantly shows between group connectivity. The total node number in each category is reduced compared to other graphs. Unlike other hive implementations nodes are shown that don't connect outside of their assigned group (eg IRG-1), but they are visualized with no prominent interconnection.
            Connection within their class (eg IRG-1 is implied by being positioned on a vertical or horizontal axis, which connects all nodes.
			"""
        )
        hive_two(first, color_code, color_code_0, reverse)

    if genre == "Bundle":
        my_expander = st.expander("show labels?")
        labels_ = my_expander.radio("Would you like to label nodes?", ("No", "Yes"))
        if labels_ == "Yes":
            labels = True
        if labels_ == "No":
            labels = False
        my_expander = st.expander("Bundling explanation")

        my_expander.markdown(
            """The graph type below is called edge bundling. "Bundling" connecting cables simplifies the visualization.\n \
			 Think of it like internet cables which are bundled. Internet backbones connect places far \n \
			 apart as to economize wiring material. Conservation of wire material is also seen in the nervous system.
             In the corpus callosum and spinal column convergent paths are constricted into relatively narrower bundles."""
        )
        with _lock:

            fig4 = data_shade(first, color_code, adj_mat, color_dict, labels)
            st.pyplot(fig4, use_column_width=True)
    if genre == "cyto":
        # from ipycytoscape import CytoscapeWidget
        # cyto = CytoscapeWidget()
        # cyto.graph.add_graph_from_networkx(first)
        # st.text(dir(cyto))
        # components.html(raw_html)
        # cyto_data = nx.cytoscape_data(first)
        # cyto_graph = nx.cytoscape_graph(cyto_data)
        # st.text(type(cyto_graph))
        # st.text(cyto_data["elements"])
        # st.text(cyto_data.keys())
        # st.text(cyto_data['directed'])
        # st.text(cyto.cytoscape_style)
        # st.text(cyto.cytoscape_layout)
        # st.text(color_dict)
        # st.text(color_code)

        G = first
        pos = nx.fruchterman_reingold_layout(G)
        A = nx.to_pandas_adjacency(first)
        nodes = [
            {
                "data": {"id": node, "label": node},
                "position": {"x": 500 * pos[node][0], "y": 500 * pos[node][1]},
                "color": color_code[node]
                #'locked': 'true'
            }
            for node in G.nodes
            if node in color_code
        ]

        edges = []
        for col in A:
            for row, value in A[col].iteritems():
                if {"data": {"source": row, "target": col}} not in edges and row != col:
                    edges.append({"data": {"source": col, "target": row}})

        for edge in edges:
            edge["data"]["weight"] = (
                0.1 * A.loc[edge["data"]["source"], edge["data"]["target"]]
            )

        elements1 = nodes + edges

        # from streamlit_cytoscapejs import st_cytoscapejs

        # elements = [{"data":cyto_data['elements']}]
        # st.text(elements)

        import streamlit_bd_cytoscapejs

        elements = elements1  # cyto_data#[{"data":cyto_data['elements']}]
        # st.text(cyto_data['elements'])
        layout = {"name": "random"}
        layout = {"name": "preset"}
        # grid
        # circle
        # concentric
        # breadthfirst
        # cose
        stylesheet = [
            {"selector": "node", "style": {"label": "data(id)"}},
            {
                "selector": "edge",
                "style": {
                    # The default curve style does not work with certain arrows
                    "curve-style": "bezier"
                },
            },
            {
                "selector": "#BA",
                "style": {
                    "source-arrow-color": "red",
                    "source-arrow-shape": "triangle",
                    "line-color": "red",
                },
            },
            {
                "selector": "#DA",
                "style": {
                    "target-arrow-color": "blue",
                    "target-arrow-shape": "vee",
                    "line-color": "blue",
                },
            },
            {
                "selector": "#BC",
                "style": {
                    "mid-source-arrow-color": "green",
                    "mid-source-arrow-shape": "diamond",
                    "mid-source-arrow-fill": "hollow",
                    "line-color": "green",
                },
            },
            {
                "selector": "#CD",
                "style": {
                    "mid-target-arrow-color": "black",
                    "mid-target-arrow-shape": "circle",
                    "arrow-scale": 2,
                    "line-color": "black",
                },
            },
        ]

        node_id = streamlit_bd_cytoscapejs.st_bd_cytoscape(
            elements, layout=layout, key="foo"
        )
        st.write(node_id)

        import dash_cytoscape as cyto
        import dash_html_components as html

        # app = dash.Dash(__name__)
        # layout = html.Div([
        cyto.Cytoscape(id="cytoscape", elements=elements, layout={"name": "preset"})
        # ])
        # st.text(dir(cyto))
        # st.write(cyto)
        # st.text()
        # st.write(layout.to_plotly_json())
        # components.html(layout.to_plotly_json())

    if genre == "Basic":
        #'plt.rcParams['legend.title_fontsize'] = 'xx-large'
        # my_expander = st.expander("Toggle Backbone")

        # backbone = my_expander.radio("Would you like to see the back-bone?", ("No", "Yes"))
        # if backbone:
        #    alpha = 0.05
        #    G = disparity_filter(first)
        #    first = nx.Graph([(u, v, d) for u, v, d in G.edges(data=True) if d['alpha'] < alpha])

        my_expander = st.expander("show labels?")

        labels_ = my_expander.radio("Would you like to label nodes?", ("No", "Yes"))
        if labels_ == "Yes":
            labels_ = True
        if labels_ == "No":
            labels_ = False
        exp = st.expander("Information about Force Directed Layout")
        exp.markdown(
            """This is probably not the most informative layout option. Contrast this force directed layout network layout with bundling (wire cost is not economized here).
            The basic force directed layout is very similar to the physics engine layout, but without interactivity.
            Qoute from wikipedia:
            'Force-directed graph drawing algorithms assign forces among the set of edges and the set of nodes of a graph drawing. Typically, spring-like attractive forces based on Hooke's law are used to attract pairs of endpoints of the graph's edges towards each other, while simultaneously repulsive forces like those of electrically charged particles based on Coulomb's law are used to separate all pairs of nodes. In equilibrium states for this system of forces, the edges tend to have uniform length (because of the spring forces), and nodes that are not connected by an edge tend to be drawn further apart (because of the electrical repulsion). Edge attraction and vertex repulsion forces may be defined using functions that are not based on the physical behavior of springs and particles; for instance, some force-directed systems use springs whose attractive force is logarithmic rather than linear.'
            \n
            https://en.wikipedia.org/wiki/Force-directed_graph_drawing \n
            What this means is conflicting forces of attraction, and repulsion determine node position.
            node centrality does not necessarily determine a central position.
            Also nodes can be central because of high in-degree out-degree or both.

            [LC Freeman A set of Measures of Betweeness (1977)](https://www.jstor.org/stable/pdf/3033543.pdf?casa_token=TzgYRJHfiYwAAAAA:r_8UKsxHRT7GRzoZ1OXwhJpzBbalTBYbG53me2fyMgZOvHnS9XM5TGB5yusfk5mCzQqXz4exAEFUcKXZ8I5ciIlU2dGpADzDfMu4Zm0rdA65G_ZzzJGo)
            """
        )
        H = first.to_undirected()

        centrality = nx.betweenness_centrality(H, k=10, endpoints=True)
        edge_thickness = [v * 20000 for v in centrality.values()]
        node_size = [v * 20000 for v in centrality.values()]

        # compute community structure
        lpc = nx.community.label_propagation_communities(H)
        community_index = {n: i for i, com in enumerate(lpc) for n in com}
        with _lock:

            #### draw graph ####
            fig, ax = plt.subplots(figsize=(20, 15))
            # fig, ax = plt.subplots(figsize=(15,15))

            pos = nx.spring_layout(H, k=0.05, seed=4572321, scale=1)

            node_color = [color_code[n] for n in H]
            srcs = list(adj_mat["src"].values)

            srcs = []
            for e in H.edges:
                src = color_code[e[0]]
                srcs.append(src)
            nx.draw_networkx_nodes(
                H,
                pos=pos,
                node_color=node_color,
                node_size=node_size,
                alpha=0.5,
                linewidths=2,
            )

            labels = {}
            for node in H.nodes():
                # set the node name as the key and the label as its value
                labels[node] = node
            if labels_:
                nx.draw_networkx_labels(H, pos, labels, font_size=16, font_color="r")

            axx = fig.gca()  # to get the current axis
            axx.collections[0].set_edgecolor("#FF0000")
            nx.draw_networkx_edges(
                H, pos=pos, edge_color=srcs, alpha=0.5, width=list(adj_mat["weight"].values)
            )

            # Title/legend
            font = {"color": "k", "fontweight": "bold", "fontsize": 20}
            ax.set_title("network", font)
            # Change font color for legend
            font["color"] = "b"

            ax.text(
                0.80,
                0.06,
                "node size = betweeness centrality",
                horizontalalignment="center",
                transform=ax.transAxes,
                fontdict=font,
            )

            # Resize figure for label readibility
            ax.margins(0.1, 0.05)
            fig.tight_layout()
            plt.axis("off")

            for k, v in color_dict.items():
                plt.scatter([], [], c=v, label=k)

            plt.legend(frameon=False, prop={"size": 24})
            # leg = ax.legend()
            # leg.set_title()
            st.pyplot(fig, use_column_width=True)

    adj_mat = pd.DataFrame(adj_mat_dicts)
    narr = nx.to_pandas_adjacency(first)

    ideo_colors = [
        "rgba(244, 109, 67, 0.75)",
        "rgba(253, 174, 97, 0.75)",
        "rgba(254, 224, 139, 0.75)",
        "rgba(217, 239, 139, 0.75)",
        "rgba(166, 217, 106, 0.75)",
        "rgba(244, 109, 67, 0.75)",
        "rgba(253, 174, 97, 0.75)",
        "rgba(254, 224, 139, 0.75)",
        "rgba(217, 239, 139, 0.75)",
        "rgba(166, 217, 106, 0.75)",
    ]

    if genre == "AdjacencyMatrix":

        my_expander = st.expander("Explanation of Adjacency Clustergrams")

        my_expander.markdown(
            """Clustergrams are re-sorted adjacency matrices, thats why their diaganols do not appear to be zero. \n
			The sorting maximizes cluster size.

            In all three graphs below, only every second node is labelled on x,y axis. This is so as not over crowd the
            axis labels. If you look closely at the pixels, pixels vary at double the frequency of the node labels.
			"""
        )
        with _lock:

            g = sns.clustermap(df2)
            st.pyplot(g)

        st.markdown(
            "un sorted interactive adjacency matrix (data not organized to emphasise clusters)"
        )
        st.markdown(
            "Diagnols are still not zero, because column names and row names are not sorted."
        )

        fig = plot_imshow_plotly(df2)
        st.write(fig)
        st.markdown("sorted interactive clustergram of adjacency matrix")

        columns = list(df2.columns.values)
        rows = list(df2.index)

        try:
            figure = dashbio.Clustergram(
                data=df2.loc[rows].values,
                column_labels=columns,
                row_labels=rows,
                color_threshold={"row": 0, "col": 70},
                hidden_labels="row",
                height=800,
                width=800,
            )

            st.write(figure)
        except:
            pass

    if genre == "Chord":
        # https://docs.bokeh.org/en/0.12.3/docs/gallery/chord_chart.html
        # from bokeh import Chord
        # nodes = data['nodes']
        # links = data['links']

        # nodes_df = pd.DataFrame(nodes)
        # links_df = pd.DataFrame(links)

        # source_data = links_df.merge(nodes_df, how='left', left_on='source', right_index=True)
        # source_data = source_data.merge(nodes_df, how='left', left_on='target', right_index=True)
        # source_data = source_data[source_data["value"] > 5]

        # chord_from_df = Chord(source_data, source="name_x", target="name_y", value="value")
        # st.markdown(""" clicking on a node highlights its direct projections""")

        H = first.to_undirected()
        # T = nx.minimum_spanning_tree(H)
        st.markdown("Betweeness Centrality:")
        st.markdown("Top to bottom node id from most central to least:")
        centrality = nx.betweenness_centrality(H, k=10, endpoints=True)

        # centrality = nx.betweenness_centrality(H)#, endpoints=True)
        df = pd.DataFrame([centrality])
        df = df.T
        df.sort_values(0, axis=0, ascending=False, inplace=True)
        bc = df
        bc.rename(columns={0: "centrality value"}, inplace=True)
        # st.write(bc.head())
        # st.markdown("In degree Centrality:")
        # st.markdown("Top to bottom node id from most central to least:")

        temp = pd.DataFrame(first.nodes)
        nodes = hv.Dataset(temp[0])

        links = copy.copy(adj_mat)
        links.rename(
            columns={"weight": "value", "src": "source", "tgt": "target"}, inplace=True
        )
        links = links[links["value"] != 0]

        Nodes_ = set(
            links["source"].unique().tolist() + links["target"].unique().tolist()
        )
        Nodes = {node: i for i, node in enumerate(Nodes_)}

        df_links = links.replace({"source": Nodes, "target": Nodes})
        for k in Nodes.keys():
            if k not in color_code_0.keys():
                color_code_0[k] = "Unknown"

        df_nodes = pd.DataFrame(
            {
                "index": [idx for idx in Nodes.values()],
                "name": [name for name in Nodes.keys()],
                "colors": [color_code_0[k] for k in Nodes.keys()],
            }
        )
        dic_to_sort = {}
        for i, kk in enumerate(df_nodes["name"]):
            dic_to_sort[i] = color_code_0[k]

        t = pd.Series(dic_to_sort)
        df_nodes["sort"] = t  # pd.Series(df_links.source)
        df_nodes.sort_values(by=["sort"], inplace=True)

        dic_to_sort = {}
        for i, kk in enumerate(df_links["source"]):
            k = df_nodes.loc[kk, "name"]
            # st.text(k)
            if k not in color_code_0.keys():
                color_code_0[k] = "Unknown"
            df_nodes.loc[kk, "colors"] = color_code_0[k]
            dic_to_sort[i] = color_code_0[k]

        pd.set_option("display.max_columns", 11)
        hv.extension("bokeh")
        hv.output(size=200)
        t = pd.Series(dic_to_sort)
        df_links["sort"] = t  # pd.Series(df_links.source)
        df_links.sort_values(by=["sort"], inplace=True)
        # df_links['colors'] = None
        categories = np.unique(df_links["sort"])
        colors = np.linspace(0, 1, len(categories))
        colordicth = dict(zip(categories, colors))

        df_links["Color"] = df_links["sort"].apply(lambda x: float(colordicth[x]))
        # for i,row in df_links.iterrows():
        #    st.text(i)
        #    if row[-1]['sort'] == "IRG 1":
        #        row[-1]
        # if row[-2] == "IRG 1":
        # if df_links.loc[i,'sort'] == "IRG 1":
        # st.text(df_links.loc[i,'sort'])

        # df_links.loc[i,'colors']
        # df_nodes["index"] = df_links["Color"]
        # st.write(df_links)
        # st.write()

        # https://geomdata.gitlab.io/hiveplotlib/karate_club.html
        # Todo make hiveplot
        #
        # st.text(chord.transform)
        # colors,y = chord.transform(chord,"Color")
        # st.text(colors)
        # st.text(dir(chord))
        # from bokeh.sampledata.les_mis import data

        # links = pd.DataFrame(data['links'])
        # nodes = hv.Dataset(pd.DataFrame(data['nodes']), 'index')
        # hv.Chord((links, nodes)).select(value=(5, None)).opts(

        # st.write(hv.render((chordt), backend="bokeh"))

        # st.text(links.head())
        # st.text(links.tail())
        # from chord3 import doCircleRibbonGraph
        # labels = first.nodes
        colors = df_links["Color"].values
        # temp = nx.to_pandas_adjacency(first)
        # temp = temp[temp!=0]
        # temp = temp[temp!=np.nan]
        # st.write(temp)
        # doCircleRibbonGraph(temp, labels, colors, plot_size=400, title="Phd Country")
        # (matrix, labels, colors, plot_size=400, title="Phd Country")
        nodes = hv.Dataset(df_nodes, "index")
        # st.write(nodes)
        # st.write(df_links)
        df_links["index"] = df_links["Color"]
        chord = hv.Chord(
            (df_links, nodes)
        )  # .opts.Chord(cmap='Category20', edge_color=dim('source').astype(str), node_color=dim('index').astype(str))
        # .select(value=(5, None))

        chord.opts(
            opts.Chord(
                cmap="Category20",
                edge_cmap="Category20",
                edge_color=dim("sort").str(),
                width=350,
                height=350,
                labels="Color",
            )
        )
        # st.markdown("Chord layout democratic")
        # sankey = hv.Sankey(df_links, label='Energy Diagram')
        # sankey.opts(label_position='left', edge_color='target', node_color='index', cmap='tab20')

        hv.save(chord, "chord2.html", backend="bokeh")
        HtmlFile2 = open("chord2.html", "r", encoding="utf-8")
        source_code2 = HtmlFile2.read()
        components.html(source_code2, height=750, width=750)

    def dontdo():

        edges_df = links.reset_index(drop=True)
        graph = hv.Graph(edges_df)
        # opts.defaults(opts.Nodes(size=5, padding=0.1))
        from holoviews.operation.datashader import (
            datashade,
            dynspread,
            directly_connect_edges,
            bundle_graph,
            stack,
        )

        # st.markdown("bundling + chord")
        # st.markdown("Able to show that not everything is connected to everything else")
        """
		circular = bundle_graph(graph)
		datashade(circular, width=500, height=500) * circular.nodes
		st.write(hv.render((circular), backend="bokeh"))
		st.markdown("clustergram of adjacency matrix: These don't look the same as sorting algorithms are different")
		"""

    # flux = np.array([[11975,  5871, 8916, 2868],
    #  [ 1951, 10048, 2060, 6171],
    #  [ 8010, 16145, 8090, 8045],
    #  [ 1013,   990,  940, 6907]
    # ])
    # flux = narr

    # ax = plt.axes([0,0,1,1])
    # from chord4 import chordDiagram
    # nodePos = chordDiagram(flux, ax, colors=[hex2rgb(x) for x in ['#666666', '#66ff66', '#ff6666', '#6666ff']])
    # nodePos = chordDiagram(flux, ax)
    # ax.axis('off')
    # prop = dict(fontsize=16*0.8, ha='center', va='center')
    # nodes = ['non-crystal', 'FCC', 'HCP', 'BCC']
    # for i in range(4):
    #    ax.text(nodePos[i][0], nodePos[i][1], nodes[i], rotation=nodePos[i][2], **prop)

    # plt.savefig("example.png", dpi=600,
    #        transparent=True,
    #        bbox_inches='tight', pad_inches=0.02)
    # st.pyplot(fig)
    # circular = bundle_graph(graph)
    # chord = hv.Chord(links)
    # datashade(chord, width=300, height=300) #* circular.nodes
    # overlay.opts(opts.Graph(edge_line_color='white', edge_hover_line_color='blue', padding=0.1))
    # st.write(hv.render((chord), backend="bokeh"))

    # st.write(chord)
    # st.write(links)
    # st.write(nodes.data.head())
    # chord = hv.Chord((links,nodes))#.select(value=(5, None))
    # chord.opts(opts.Chord(cmap='Category20',
    # 		   edge_cmap='Category20', edge_color=dim('source').str(),
    #           labels='name', node_color=node_color))
    # st.write(hv.render((chord), backend="bokeh"))
    # st.hvplot(chord)
    # st.write(hv.render(chord, backend="bokeh"))

    # st.write(adj_mat3)
    # chord3 = chord2.make_filled_chord(adj_mat3)
    # st.write(chord3)

    # nodes = first.nodes
    # edges = first.edges
    # value_dim = [i*10 for i in adj_mat["weight"]]
    # careers = hv.Sankey((edges, nodes), ['From', 'To'])#, vdims=value_dim)

    # careers.opts(
    #    opts.Sankey(labels='label', label_position='right', width=900, height=300, cmap='Set1',
    #                edge_color=dim('To').str(), node_color=dim('index').str()))
    # careers.write(hv.render(graph, backend="bokeh"))

    # for trace in edge_trace:
    # 	fig.add_trace(trace)  # Add node trace
    # fig.add_trace(node_trace)  # Remove legend

    # fig.show()


if __name__ == "__main__":
    # import os
    # os.system("python make_serial_plots0.py")
    # os.system("python make_serial_plots1.py")

    main()


def dontdo():

    link = dict(source=adj_mat["src"], target=adj_mat["tgt"], value=adj_mat["weight"])

    # generate_sankey_figure(list(first.nodes), adj_mat,title = 'Sankey Diagram')

    fig = go.Figure(
        data=[
            go.Sankey(
                node=dict(
                    pad=15,
                    thickness=20,
                    line=dict(color="black", width=0.5),
                    label=list(first.nodes()),  # ["A1", "A2", "B1", "B2", "C1", "C2"],
                    color="blue",
                ),
                link=dict(
                    source=adj_mat["src"],
                    target=adj_mat["tgt"],
                    value=[i * 10 for i in adj_mat["weight"]],
                ),
            )
        ]
    )

    fig.update_layout(title_text="Basic Sankey Diagram", font_size=10)

    st.write(fig)
    link = dict(
        source=adj_mat["src"],
        target=adj_mat["tgt"],
        value=[i * 10 for i in adj_mat["weight"]],
    )

    data = go.Sankey(link=link)

    # fig3 = go.Figure(data)
    layout = go.Layout(
        paper_bgcolor="rgba(0,0,0,0)",  # transparent background
        plot_bgcolor="rgba(0,0,0,0)",  # transparent 2nd background
        xaxis={"showgrid": False, "zeroline": False},  # no gridlines
        yaxis={"showgrid": False, "zeroline": False},  # no gridlines
    )  # Create figure
    layout["width"] = 925
    layout["height"] = 925

    fig3 = go.Figure(data, layout=layout)  # Add all edge traces
    st.write(fig3)
